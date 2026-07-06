"""把雷达累积数据导出为带格式的Excel报表（业务侧交付形态）。

三个工作表：
- 概览：KPI块、技能需求Top10（含数据条）、岗位类别分布、技能柱状图
- 岗位明细：全部追踪岗位，冻结首行+自动筛选+隔行着色
- 地区x类别：交叉计数表

设计定位：演示「用代码自动产Excel报表」的工作形态。报表随每日管线自动重建，
数据与看板同源（jobs.csv / daily_stats.json）。

用法：python src/export_excel.py（先跑parse_skills.py）
产出：reports/数据岗位周报.xlsx
"""
from __future__ import annotations

import csv
import json
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import DataBarRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parents[1]
STATS_JSON = PROJECT_ROOT / "data" / "daily_stats.json"
JOBS_CSV = PROJECT_ROOT / "data" / "jobs.csv"
OUT = PROJECT_ROOT / "reports" / "数据岗位周报.xlsx"

FONT = "Microsoft YaHei"
DARK = "1F4E79"
LIGHT_FILL = PatternFill("solid", fgColor="EDF2F9")
HEADER_FILL = PatternFill("solid", fgColor=DARK)
HEADER_FONT = Font(name=FONT, size=10, bold=True, color="FFFFFF")
TITLE_FONT = Font(name=FONT, size=16, bold=True, color=DARK)
KPI_FONT = Font(name=FONT, size=22, bold=True, color=DARK)
LABEL_FONT = Font(name=FONT, size=9, color="6B7688")
BODY_FONT = Font(name=FONT, size=10)
THIN = Border(*[Side(style="thin", color="D8DEE8")] * 4)


def style_header_row(ws, row: int, cols: int) -> None:
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN
        cell.alignment = Alignment(horizontal="center", vertical="center")


def sheet_overview(wb: Workbook, day: dict, date: str, tracked: int) -> None:
    ws = wb.active
    ws.title = "概览"
    ws.sheet_view.showGridLines = False
    ws["B2"] = "数据岗位周报 · 数据岗位需求雷达"
    ws["B2"].font = TITLE_FONT
    ws["B3"] = (f"数据日期 {date} · 样本为7个公开API的国际远程数据岗 · "
                f"报表由export_excel.py自动生成于 "
                f"{datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M')} UTC")
    ws["B3"].font = LABEL_FONT

    kpis = [("今日在架数据岗", day["total_data_jobs"]),
            ("今日新上架", day["new_today"]),
            ("累计追踪岗位", tracked),
            ("数据源正常", f"{7 - len(day.get('failed_sources', []))}/7")]
    for i, (label, value) in enumerate(kpis):
        col = 2 + i * 2
        lc, vc = ws.cell(row=5, column=col), ws.cell(row=6, column=col)
        lc.value, lc.font = label, LABEL_FONT
        vc.value, vc.font = value, KPI_FONT

    ws["B9"] = "技能需求Top10"
    ws["B9"].font = Font(name=FONT, size=12, bold=True, color=DARK)
    headers = ["技能", "岗位数", "提及率"]
    for i, h in enumerate(headers):
        ws.cell(row=10, column=2 + i, value=h)
    style_header_row_range(ws, 10, 2, 4)
    total = day["total_data_jobs"] or 1
    skills = list(day["by_skill"].items())[:10]
    for r, (name, count) in enumerate(skills, start=11):
        ws.cell(row=r, column=2, value=name).font = BODY_FONT
        ws.cell(row=r, column=3, value=count).font = BODY_FONT
        pct = ws.cell(row=r, column=4, value=count / total)
        pct.number_format = "0.0%"
        pct.font = BODY_FONT
        for c in (2, 3, 4):
            ws.cell(row=r, column=c).border = THIN
    last = 10 + len(skills)
    ws.conditional_formatting.add(
        f"C11:C{last}",
        DataBarRule(start_type="num", start_value=0, end_type="max",
                    color=DARK, showValue=True))

    ws["F9"] = "岗位类别分布"
    ws["F9"].font = Font(name=FONT, size=12, bold=True, color=DARK)
    for i, h in enumerate(["类别", "数量", "占比"]):
        ws.cell(row=10, column=6 + i, value=h)
    style_header_row_range(ws, 10, 6, 8)
    for r, (name, count) in enumerate(day["by_category"].items(), start=11):
        ws.cell(row=r, column=6, value=name).font = BODY_FONT
        ws.cell(row=r, column=7, value=count).font = BODY_FONT
        pc = ws.cell(row=r, column=8, value=count / total)
        pc.number_format = "0.0%"
        pc.font = BODY_FONT
        for c in (6, 7, 8):
            ws.cell(row=r, column=c).border = THIN

    chart = BarChart()
    chart.type = "bar"
    chart.title = "技能需求Top10（岗位数）"
    chart.height, chart.width = 8.5, 14
    chart.legend = None
    data = Reference(ws, min_col=3, min_row=10, max_row=last)
    cats = Reference(ws, min_col=2, min_row=11, max_row=last)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "F17")

    for col, width in {"A": 2, "B": 14, "C": 9, "D": 9, "E": 3,
                       "F": 16, "G": 9, "H": 9, "I": 12, "J": 12}.items():
        ws.column_dimensions[col].width = width


def style_header_row_range(ws, row: int, col_start: int, col_end: int) -> None:
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN
        cell.alignment = Alignment(horizontal="center", vertical="center")


def sheet_details(wb: Workbook, rows: list[dict]) -> None:
    ws = wb.create_sheet("岗位明细")
    headers = ["岗位名称", "公司", "类别", "地区", "命中技能", "来源", "首次收录", "链接"]
    keys = ["title", "company", "category", "region", "skills", "source", "first_seen", "url"]
    cat_labels = {"analyst": "数据分析/BI", "scientist": "数据科学/算法",
                  "engineer": "数据开发/工程", "other_data": "其他数据岗"}
    ws.append(headers)
    style_header_row(ws, 1, len(headers))
    for i, row in enumerate(rows):
        vals = []
        for k in keys:
            v = row.get(k, "")
            if k == "category":
                v = cat_labels.get(v, v)
            if k == "skills":
                v = (v or "").replace("|", "、")
            vals.append(v)
        ws.append(vals)
        r = i + 2
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = BODY_FONT
            cell.border = THIN
            if i % 2 == 1:
                cell.fill = LIGHT_FILL
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"
    for col, width in {"A": 42, "B": 20, "C": 13, "D": 10, "E": 30,
                       "F": 13, "G": 11, "H": 45}.items():
        ws.column_dimensions[col].width = width


def sheet_crosstab(wb: Workbook, rows: list[dict]) -> None:
    ws = wb.create_sheet("地区x类别")
    cat_labels = {"analyst": "数据分析/BI", "scientist": "数据科学/算法",
                  "engineer": "数据开发/工程", "other_data": "其他数据岗"}
    regions = sorted({r["region"] for r in rows})
    cats = list(cat_labels.values())
    counts = {(rg, ct): 0 for rg in regions for ct in cats}
    for r in rows:
        key = (r["region"], cat_labels.get(r["category"], r["category"]))
        if key in counts:
            counts[key] += 1
    ws.append(["地区"] + cats + ["合计"])
    style_header_row(ws, 1, len(cats) + 2)
    for i, rg in enumerate(regions):
        vals = [counts[(rg, ct)] for ct in cats]
        ws.append([rg] + vals + [sum(vals)])
        for c in range(1, len(cats) + 3):
            cell = ws.cell(row=i + 2, column=c)
            cell.font = BODY_FONT
            cell.border = THIN
    ws.column_dimensions["A"].width = 14
    for c in range(2, len(cats) + 3):
        ws.column_dimensions[get_column_letter(c)].width = 13


def main() -> None:
    stats = json.loads(STATS_JSON.read_text(encoding="utf-8"))
    days = stats["days"]
    date = sorted(days.keys())[-1]
    with open(JOBS_CSV, encoding="utf-8-sig", newline="") as f:
        rows = list(csv.DictReader(f))

    wb = Workbook()
    sheet_overview(wb, days[date], date, len(rows))
    sheet_details(wb, rows)
    sheet_crosstab(wb, rows)
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Excel报表已生成: {OUT} （{len(rows)}条岗位明细）")


if __name__ == "__main__":
    main()
